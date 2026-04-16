import os
import re
import pandas as pd
import numpy as np
from openpyxl import load_workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

root_dir = 'model_sonucları'

def parse_file(scores):
    metrics = re.findall(
        r'nochange_metric:\n({.*?})nochange_acc:([\d.]+)\nchange_metric:\n({.*?})change_acc:([\d.]+)\n.*\n({.*?})',
        scores, re.DOTALL)
    nochange = eval(metrics[0][0])
    nochange_acc = float(metrics[0][1])
    change = eval(metrics[0][2])
    change_acc = float(metrics[0][3])
    overall = eval(metrics[0][4])
    return nochange, nochange_acc, change, change_acc, overall

# Metric sets for computing S_m^*
avg_metrics_change = ["Bleu_4", "ROUGE_L", "METEOR", "CIDEr", "SPICE"]
avg_metrics_nochange = ["Bleu_4", "ROUGE_L", "METEOR", "SPICE"]

def prepare_dataframe(rows, index_names, mode):
    df = pd.DataFrame(rows, index=index_names)
    df.index.name = 'Folder'

    # Multiply by 100 and round to 4 decimals
    df = (df * 100).round(4)

    # Determine which metrics to average
    if mode == "No Change":
        relevant_metrics = [m for m in avg_metrics_nochange if m in df.columns]
    else:
        relevant_metrics = [m for m in avg_metrics_change if m in df.columns]

    # Add S_m^* column
    df["$S_m^*$"] = df[relevant_metrics].mean(axis=1).round(4)

    return df

# Collect metrics
change_rows, nochange_rows, overall_rows = [], [], []
index_names = []

for subdir, _, files in os.walk(root_dir):
    for file in files:
        if file.endswith('.txt') and 'test_output' in file:
            file_path = os.path.join(subdir, file)
            try:
                with open(file_path, 'r') as f:
                    scores = f.read()

                nochange, _, change, _, overall = parse_file(scores)
                folder_name = os.path.basename(os.path.dirname(file_path))

                change_rows.append(change)
                nochange_rows.append(nochange)
                overall_rows.append(overall)
                index_names.append(folder_name)

            except Exception as e:
                print(f"❌ Error parsing {file_path}: {e}")

# Prepare final DataFrames
df_change = prepare_dataframe(change_rows, index_names, "Change")
df_nochange = prepare_dataframe(nochange_rows, index_names, "No Change")
df_overall = prepare_dataframe(overall_rows, index_names, "Overall")

# Save to Excel
output_file = "metrics_grouped_sheets.xlsx"
with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
    df_change.to_excel(writer, sheet_name='Change')
    df_nochange.to_excel(writer, sheet_name='No Change')
    df_overall.to_excel(writer, sheet_name='Overall')

# Bold max values per column (except index column)
def bold_max_in_sheet(sheet):
    for col in range(2, sheet.max_column + 1):
        max_val = None
        max_row = None
        for row in range(2, sheet.max_row + 1):
            val = sheet.cell(row=row, column=col).value
            if isinstance(val, (int, float)):
                if (max_val is None) or (val > max_val):
                    max_val = val
                    max_row = row
        if max_row:
            sheet.cell(row=max_row, column=col).font = Font(bold=True)

# Reopen workbook and apply formatting
wb = load_workbook(output_file)
for sheetname in ['Change', 'No Change', 'Overall']:
    bold_max_in_sheet(wb[sheetname])
wb.save(output_file)

print(f"✅ Saved to '{output_file}' with bolded maximums and $S_m^*$ column.")