import torch
import torch.optim
import torch.utils.data
import torchvision.transforms as transforms
from datasets import CaptionDataset
from utils import get_eval_score, get_eval_score_detailed
import torch.nn.functional as F
from tqdm import tqdm
import argparse
import time
import os
import json
import numpy as np
from glob import glob
from openpyxl import Workbook

# --- Globals and Setup (Matching eval_changed.py) ---
device = torch.device("cuda" if torch.cuda.is_available() else "cpu")

normalize = transforms.Normalize(mean=[0.485, 0.456, 0.406],
                                 std=[0.229, 0.224, 0.225])

CATEGORY_ORDER = ["WF", "FL", "WET", "GL", "AG", "GF", "NO", "UR"]


# Helper to match legacy get_key behavior
def get_key(dict_, value):
    return [k for k, v in dict_.items() if v == value]


def derive_run_id(ckpt_path):
    # Extracts a readable ID from the path, e.g., "20260107_0"
    try:
        parts = ckpt_path.replace("\\", "/").split("/")
        # usually .../20260107_0/model_dir/checkpoint.pth
        return parts[-3]
    except:
        return "unknown_run"

def metric_to_pct(x):
    if x is None:
        return None
    return float(x) * 100.0

def compute_sm_star(metric_dict, category=None):
    """
    Category-aware S_m^* on 100-scaled metrics.

    Default:
        S_m^* = average of BLEU-4, METEOR, ROUGE_L, CIDEr, SPICE

    For NO:
        exclude CIDEr
        S_m^* = average of BLEU-4, METEOR, ROUGE_L, SPICE
    """
    b4 = metric_to_pct(metric_dict.get("Bleu_4", 0.0))
    met = metric_to_pct(metric_dict.get("METEOR", 0.0))
    rouge = metric_to_pct(metric_dict.get("ROUGE_L", 0.0))
    spice = metric_to_pct(metric_dict.get("SPICE", 0.0))

    if category == "NO":
        return (b4 + met + rouge + spice) / 4.0

    cider = metric_to_pct(metric_dict.get("CIDEr", 0.0))
    return (b4 + met + rouge + cider + spice) / 5.0


def evaluate_checkpoint_exact(args, ckpt_path):
    print(f"\n[INFO] Evaluating: {ckpt_path}")

    # --- 1. Load Model (Legacy Mode) ---
    checkpoint = torch.load(ckpt_path, map_location=device, weights_only=False)
    encoder_image = checkpoint['encoder_image'].to(device).eval()
    encoder_feat = checkpoint['encoder_feat'].to(device).eval()
    decoder = checkpoint['decoder'].to(device).eval()

    # --- 2. Load Word Map ---
    word_map_file = os.path.join(args.data_folder, 'WORDMAP_' + args.data_name + '.json')
    with open(word_map_file, 'r') as f:
        word_map = json.load(f)
    vocab_size = len(word_map)

    # --- 3. DataLoader ---
    # Shuffle=False is critical to match indices
    dataset = CaptionDataset(args.data_folder, args.data_name, args.Split, transform=transforms.Compose([normalize]))

    # --- CRITICAL ADAPTATION FOR INDEX TRACKING ---
    # We retrieve cpi and the filtered image_names directly from the dataset instance.
    cpi = dataset.cpi
    image_names = dataset.image_names

    # [FIX] Retrieve the mapping to original indices to prevent mismatches
    orig_indices = dataset.orig_img_indices

    loader = torch.utils.data.DataLoader(
        dataset,
        batch_size=1, shuffle=False, num_workers=0, pin_memory=False
    )

    # --- 4. Containers for Logic ---
    references = []
    hypotheses = []

    # For console printout (Aggregated stats)
    change_references = []
    change_hypotheses = []
    nochange_references = []
    nochange_hypotheses = []
    change_acc = 0
    nochange_acc = 0

    # For Per-Sample JSON
    sample_metadata = []  # Stores index, name, category, text refs

    # --- NEW: category-level corpus containers ---
    category_references = {cat: [] for cat in CATEGORY_ORDER}
    category_hypotheses = {cat: [] for cat in CATEGORY_ORDER}

    nochange_list = [
        "the scene is the same as before ",
        "there is no difference ",
        "two scenes seem identical ",
        "no change has occurred ",
        "almost nothing has changed ",
    ]

    beam_size = args.beam_size

    with torch.no_grad():
        for i, (image_pairs, caps, caplens, allcaps) in enumerate(
                tqdm(loader, desc=f"EVAL {derive_run_id(ckpt_path)}")):

            # --- EXACT Legacy Logic Adapted for Dynamic CPI ---
            # Use cpi from dataset instead of hardcoded 5
            if (i + 1) % cpi != 0:
                continue

            # Identify Image Metadata
            # i is the index in the full filtered caption list.
            # (i+1) is a multiple of cpi.
            # integer division gives the index in the dataset.image_names list
            img_idx_dataset = (i // cpi)

            if img_idx_dataset >= len(image_names):
                break

            img_name = image_names[img_idx_dataset]

            # [FIX] Use the original index from the dataset mapping, NOT the loop counter
            real_index = orig_indices[img_idx_dataset]

            # Category extraction based on file naming convention (e.g., WF_704 -> WF)
            category = img_name.split('_')[0] if '_' in img_name else "UNK"

            # --- Start Beam Search (Verbatim from eval_changed.py) ---
            k = beam_size
            Caption_End = False
            image_pairs["before"] = image_pairs["before"].to(device)
            image_pairs["after"] = image_pairs["after"].to(device)

            # Encode MS-only
            feat_before_ms = encoder_image(image_pairs["before"])
            feat_after_ms = encoder_image(image_pairs["after"])

            imgs_A = feat_before_ms
            imgs_B = feat_after_ms
            encoder_out = encoder_feat(imgs_A, imgs_B)

            tgt = torch.zeros(52, k).to(device).to(torch.int64)
            tgt_length = tgt.size(0)
            mask = (torch.triu(torch.ones(tgt_length, tgt_length)) == 1).transpose(0, 1)
            mask = mask.float().masked_fill(mask == 0, float('-inf')).masked_fill(mask == 1, float(0.0))
            mask = mask.to(device)

            tgt[0, :] = torch.LongTensor([word_map['<start>']] * k).to(device)
            seqs = torch.LongTensor([[word_map['<start>']] * 1] * k).to(device)
            top_k_scores = torch.zeros(k, 1).to(device)
            complete_seqs = []
            complete_seqs_scores = []
            step = 1

            k_prev_words = tgt.permute(1, 0)
            S = encoder_out.size(0)
            encoder_dim = encoder_out.size(-1)

            encoder_out = encoder_out.expand(S, k, encoder_dim)
            encoder_out = encoder_out.permute(1, 0, 2)

            while True:
                tgt = k_prev_words.permute(1, 0)
                tgt_embedding = decoder.vocab_embedding(tgt)
                tgt_embedding = decoder.position_encoding(tgt_embedding)

                encoder_out = encoder_out.permute(1, 0, 2)
                pred = decoder.transformer(tgt_embedding, encoder_out, tgt_mask=mask)
                encoder_out = encoder_out.permute(1, 0, 2)

                pred = decoder.wdc(pred)
                scores = pred.permute(1, 0, 2)
                scores = scores[:, step - 1, :].squeeze(1)
                scores = F.log_softmax(scores, dim=1)

                scores = top_k_scores.expand_as(scores) + scores

                if step == 1:
                    top_k_scores, top_k_words = scores[0].topk(k, 0, True, True)
                else:
                    top_k_scores, top_k_words = scores.view(-1).topk(k, 0, True, True)

                prev_word_inds = torch.div(top_k_words, vocab_size, rounding_mode='floor')
                next_word_inds = top_k_words % vocab_size

                seqs = torch.cat([seqs[prev_word_inds], next_word_inds.unsqueeze(1)], dim=1)

                incomplete_inds = [ind for ind, next_word in enumerate(next_word_inds) if
                                   next_word != word_map['<end>']]
                complete_inds = list(set(range(len(next_word_inds))) - set(incomplete_inds))

                if len(complete_inds) > 0:
                    Caption_End = True
                    complete_seqs.extend(seqs[complete_inds].tolist())
                    complete_seqs_scores.extend(top_k_scores[complete_inds])

                k -= len(complete_inds)
                if k == 0:
                    break

                seqs = seqs[incomplete_inds]
                encoder_out = encoder_out[prev_word_inds[incomplete_inds]]
                top_k_scores = top_k_scores[incomplete_inds].unsqueeze(1)

                k_prev_words = k_prev_words[incomplete_inds]
                k_prev_words[:, :step + 1] = seqs

                if step > 50:
                    break
                step += 1

            if len(complete_seqs_scores) == 0:
                complete_seqs.extend(seqs[complete_inds].tolist())
                complete_seqs_scores.extend(top_k_scores[complete_inds])

            if len(complete_seqs_scores) > 0:
                assert Caption_End
                indices = complete_seqs_scores.index(max(complete_seqs_scores))
                seq = complete_seqs[indices]

                # References extraction
                img_caps = allcaps[0].tolist()
                img_captions = list(
                    map(lambda c: [w for w in c if
                                   w not in {word_map['<start>'], word_map['<end>'], word_map['<pad>']}],
                        img_caps))

                references.append(img_captions)

                # Hypothesis extraction
                new_sent = [w for w in seq if w not in {word_map['<start>'], word_map['<end>'], word_map['<pad>']}]
                hypotheses.append(new_sent)

                # --- NEW: store for category-level corpus metrics ---
                if category in category_references:
                    category_references[category].append(img_captions)
                    category_hypotheses[category].append(new_sent)
                else:
                    # If any unexpected prefix appears, still keep it
                    if category not in category_references:
                        category_references[category] = []
                        category_hypotheses[category] = []
                    category_references[category].append(img_captions)
                    category_hypotheses[category].append(new_sent)

                # --- Strings for JSON and No-Change Logic ---

                # Reconstruct Reference Strings
                ref_strings = []
                for ref_c in img_captions:
                    s = ""
                    for wid in ref_c:
                        w_list = get_key(word_map, wid)
                        if w_list:
                            s += w_list[0] + " "
                    ref_strings.append(s.strip())

                # Reconstruct Hypothesis String
                hyp_string = ""
                hyp_line_repo = ""
                for wid in new_sent:
                    w_list = get_key(word_map, wid)
                    if w_list:
                        hyp_string += w_list[0] + " "
                        hyp_line_repo += w_list[0] + " "

                hyp_string = hyp_string.strip()

                # --- Exact Legacy Logic for Change/NoChange ---
                # It compares the 2nd reference (index 1) to check if the image is "no change"
                ref_sentence_legacy = img_captions[1]
                ref_line_repo = ""
                for ref_word_idx in ref_sentence_legacy:
                    ref_word = get_key(word_map, ref_word_idx)
                    ref_line_repo += ref_word[0] + " "

                # Classification
                if ref_line_repo not in nochange_list:
                    # CHANGE IMAGE
                    change_references.append(img_captions)
                    change_hypotheses.append(new_sent)
                    if hyp_line_repo not in nochange_list:
                        change_acc += 1
                else:
                    # NO CHANGE IMAGE
                    nochange_references.append(img_captions)
                    nochange_hypotheses.append(new_sent)
                    if hyp_line_repo in nochange_list:
                        nochange_acc += 1

                # Store metadata for JSON
                sample_metadata.append({
                    "index": real_index,  # [FIX] Use real original index
                    "image_name": img_name,
                    "category": category,
                    "hypothesis": hyp_string,
                    "refs": ref_strings
                })

    # --- 5. Console Output (Legacy Style) ---
    print(f"len(nochange_references): {len(nochange_references)}")
    print(f"len(change_references): {len(change_references)}")

    nochange_metric = {}
    change_metric = {}

    if len(nochange_references) > 0:
        print('nochange_metric:')
        nochange_metric = get_eval_score(nochange_references, nochange_hypotheses, word_map)
        print("nochange_acc:" + str(nochange_acc / len(nochange_references)))

    if len(change_references) > 0:
        print('change_metric:')
        change_metric = get_eval_score(change_references, change_hypotheses, word_map)
        print("change_acc:" + str(change_acc / len(change_references)))

    print("............................................._..........")
    metrics = get_eval_score(references, hypotheses, word_map)
    print("vocabsize: " + str(vocab_size))

    # --- NEW: Category-level corpus metrics ---
    print("Generating category-level corpus metrics...")
    category_metrics_output = []

    ordered_categories = [c for c in CATEGORY_ORDER if c in category_references] + \
                         [c for c in sorted(category_references.keys()) if c not in CATEGORY_ORDER]

    for category in ordered_categories:
        cat_refs = category_references.get(category, [])
        cat_hyps = category_hypotheses.get(category, [])

        if len(cat_refs) == 0:
            entry = {
                "category": category,
                "n_samples": 0,
                "Bleu_1": None,
                "Bleu_2": None,
                "Bleu_3": None,
                "Bleu_4": None,
                "METEOR": None,
                "ROUGE_L": None,
                "CIDEr": None,
                "SPICE": None,
                "S_m_star": None
            }
        else:
            cat_metrics = get_eval_score(cat_refs, cat_hyps, word_map)
            sm_star = compute_sm_star(cat_metrics, category=category)

            entry = {
                "category": category,
                "n_samples": len(cat_refs),
                "Bleu_1": metric_to_pct(cat_metrics.get("Bleu_1", 0.0)),
                "Bleu_2": metric_to_pct(cat_metrics.get("Bleu_2", 0.0)),
                "Bleu_3": metric_to_pct(cat_metrics.get("Bleu_3", 0.0)),
                "Bleu_4": metric_to_pct(cat_metrics.get("Bleu_4", 0.0)),
                "METEOR": metric_to_pct(cat_metrics.get("METEOR", 0.0)),
                "ROUGE_L": metric_to_pct(cat_metrics.get("ROUGE_L", 0.0)),
                "CIDEr": metric_to_pct(cat_metrics.get("CIDEr", 0.0)),
                "SPICE": metric_to_pct(cat_metrics.get("SPICE", 0.0)),
                "S_m_star": float(sm_star)
            }

        category_metrics_output.append(entry)

    # --- 6. Generate Detailed Per-Sample JSON ---
    print("Generating per-sample detailed metrics...")

    # We use get_eval_score_detailed to get scores aligned with indices
    _, per_sample_scores = get_eval_score_detailed(references, hypotheses, word_map, return_per_sample=True,
                                                   verbose=False)

    final_json_output = []

    for i, meta in enumerate(sample_metadata):
        scores = per_sample_scores[i]

        # Calculate AVG: (B4 + Met + Rouge + Cider) / 4
        # Note: preserved as-is to avoid changing unrelated per-sample logic
        b4 = metric_to_pct(scores.get("Bleu_4", 0.0))
        met = metric_to_pct(scores.get("METEOR", 0.0))
        rouge = metric_to_pct(scores.get("ROUGE_L", 0.0))
        cider = metric_to_pct(scores.get("CIDEr", 0.0))

        avg_score = (b4 + met + rouge + cider) / 4.0

        entry = {
            "index": meta["index"],
            "image_name": meta["image_name"],
            "category": meta["category"],
            "hypothesis": meta["hypothesis"],
            # Flatten refs 1-5
            "ref_1": meta["refs"][0] if len(meta["refs"]) > 0 else "",
            "ref_2": meta["refs"][1] if len(meta["refs"]) > 1 else "",
            "ref_3": meta["refs"][2] if len(meta["refs"]) > 2 else "",
            "ref_4": meta["refs"][3] if len(meta["refs"]) > 3 else "",
            "ref_5": meta["refs"][4] if len(meta["refs"]) > 4 else "",
            # Metrics
            "Bleu_1": metric_to_pct(scores.get("Bleu_1", 0.0)),
            "Bleu_2": metric_to_pct(scores.get("Bleu_2", 0.0)),
            "Bleu_3": metric_to_pct(scores.get("Bleu_3", 0.0)),
            "Bleu_4": b4,
            "METEOR": met,
            "ROUGE_L": rouge,
            "CIDEr": cider,
            "AVG": avg_score
        }
        final_json_output.append(entry)

    return final_json_output, category_metrics_output

def safe_sheet_name(name: str) -> str:
    invalid = ['\\', '/', '*', '?', ':', '[', ']']
    for ch in invalid:
        name = name.replace(ch, "_")
    return name[:31]


def ordered_category_rows(rows):
    row_map = {str(r.get("category", "")): r for r in rows}
    ordered = []
    for cat in CATEGORY_ORDER:
        if cat in row_map:
            ordered.append(row_map[cat])
    for cat, row in row_map.items():
        if cat not in CATEGORY_ORDER:
            ordered.append(row)
    return ordered


def write_category_rows_to_sheet(ws, rows):
    headers = [
        "category", "n_samples",
        "Bleu_1", "Bleu_2", "Bleu_3", "Bleu_4",
        "METEOR", "ROUGE_L", "CIDEr", "SPICE", "S_m_star"
    ]
    ws.append(headers)

    for row in ordered_category_rows(rows):
        ws.append([
            row.get("category"),
            row.get("n_samples"),
            row.get("Bleu_1"),
            row.get("Bleu_2"),
            row.get("Bleu_3"),
            row.get("Bleu_4"),
            row.get("METEOR"),
            row.get("ROUGE_L"),
            row.get("CIDEr"),
            row.get("SPICE"),
            row.get("S_m_star"),
        ])


def build_category_excel_report(out_dir, split_name):
    json_paths = sorted(glob(os.path.join(out_dir, f"category_metrics_*_{split_name}.json")))
    if len(json_paths) == 0:
        print("[WARNING] No category_metrics JSON files found for Excel export.")
        return

    wb = Workbook()
    default_ws = wb.active
    wb.remove(default_ws)

    all_runs = []
    for path in json_paths:
        with open(path, "r", encoding="utf-8") as f:
            rows = json.load(f)

        run_name = os.path.splitext(os.path.basename(path))[0]
        run_name = run_name.replace(f"category_metrics_", "")
        run_name = run_name.replace(f"_{split_name}", "")
        ws = wb.create_sheet(title=safe_sheet_name(run_name))
        write_category_rows_to_sheet(ws, rows)
        all_runs.append(rows)

    # --- Overall averages across all JSON files ---
    overall_ws = wb.create_sheet(title="Overall")

    headers = [
        "category", "n_samples",
        "Bleu_1", "Bleu_2", "Bleu_3", "Bleu_4",
        "METEOR", "ROUGE_L", "CIDEr", "SPICE", "S_m_star"
    ]
    overall_ws.append(headers)

    metrics_keys = [
        "Bleu_1", "Bleu_2", "Bleu_3", "Bleu_4",
        "METEOR", "ROUGE_L", "CIDEr", "SPICE", "S_m_star"
    ]

    # category -> list of rows from different runs
    category_to_rows = {}
    for run_rows in all_runs:
        for row in run_rows:
            cat = str(row.get("category", ""))
            category_to_rows.setdefault(cat, []).append(row)

    ordered_categories = [c for c in CATEGORY_ORDER if c in category_to_rows] + \
                         [c for c in sorted(category_to_rows.keys()) if c not in CATEGORY_ORDER]

    for cat in ordered_categories:
        rows = category_to_rows[cat]

        # n_samples should normally be identical across runs; mean is safe
        n_samples_vals = [float(r["n_samples"]) for r in rows if r.get("n_samples") is not None]
        n_samples_avg = sum(n_samples_vals) / len(n_samples_vals) if len(n_samples_vals) > 0 else None

        avg_metrics = {}
        for key in metrics_keys:
            vals = [float(r[key]) for r in rows if r.get(key) is not None]
            avg_metrics[key] = (sum(vals) / len(vals)) if len(vals) > 0 else None

        overall_ws.append([
            cat,
            int(round(n_samples_avg)) if n_samples_avg is not None else None,
            avg_metrics["Bleu_1"],
            avg_metrics["Bleu_2"],
            avg_metrics["Bleu_3"],
            avg_metrics["Bleu_4"],
            avg_metrics["METEOR"],
            avg_metrics["ROUGE_L"],
            avg_metrics["CIDEr"],
            avg_metrics["SPICE"],
            avg_metrics["S_m_star"],
        ])

    excel_path = os.path.join(out_dir, f"category_metrics_summary_{split_name}.xlsx")
    wb.save(excel_path)
    print(f"[SUCCESS] Saved Excel summary to: {excel_path}")

def main():
    parser = argparse.ArgumentParser(description='Detailed Evaluation')

    # --- Args strictly matching your request ---
    parser.add_argument('--data_folder', default="../MOSAIC-SEN2-CC/",
                        help='folder with data files saved by create_input_files.py.')
    parser.add_argument('--data_name', default="LEVIR_CC_5_cap_per_img_5_min_word_freq",
                        help='base name shared by data files.')

    # Model architecture args needed for loading logic if not fully pickled
    parser.add_argument('--encoder_image', default="resnet101")
    parser.add_argument('--encoder_feat', default="MCCFormers_diff_as_Q")
    parser.add_argument('--decoder', default="trans", help="decoder img2txt")

    parser.add_argument('--Split', default="VAL", help='which split to evaluate')
    parser.add_argument('--beam_size', type=int, default=5, help='beam_size.')

    parser.add_argument('--out_dir', default="./sample_eval_results_categorical", help="Folder to save per_sample json")

    # The specific checkpoint list you requested
    parser.add_argument('--checkpoints', type=str, nargs='*',
                        default=[
                            "./model_sonucları/20260102_0_val_test_switch/model_dir/NEW_BEST_checkpoint_resnet101_MCCFormers_diff_as_Q_trans.pth.tar",
                            "./model_sonucları/20260102_1_val_test_switch/model_dir/NEW_BEST_checkpoint_resnet101_MCCFormers_diff_as_Q_trans.pth.tar",
                            "./model_sonucları/20260102_2_val_test_switch/model_dir/NEW_BEST_checkpoint_resnet101_MCCFormers_diff_as_Q_trans.pth.tar", ],
                        help='List of checkpoint files.')

    args = parser.parse_args()

    # Ensure output directory exists
    os.makedirs(args.out_dir, exist_ok=True)

    # Iterate over checkpoints
    for ckpt in args.checkpoints:
        if not os.path.exists(ckpt):
            print(f"[WARNING] Checkpoint not found: {ckpt}")
            continue

        run_id = derive_run_id(ckpt)

        # Run evaluation
        results_json, category_metrics_json = evaluate_checkpoint_exact(args, ckpt)

        # Save per-sample JSON
        out_filename = f"per_sample_{run_id}_{args.Split}.json"
        out_path = os.path.join(args.out_dir, out_filename)

        with open(out_path, "w", encoding="utf-8") as f:
            json.dump(results_json, f, indent=2, ensure_ascii=False)

        # Save category-level corpus metrics JSON
        cat_out_filename = f"category_metrics_{run_id}_{args.Split}.json"
        cat_out_path = os.path.join(args.out_dir, cat_out_filename)

        with open(cat_out_path, "w", encoding="utf-8") as f:
            json.dump(category_metrics_json, f, indent=2, ensure_ascii=False)

        print(f"[SUCCESS] Saved per-sample results to: {out_path}")
        print(f"[SUCCESS] Saved category metrics to: {cat_out_path}\n" + "=" * 50)
    # Build one Excel workbook from all category_metrics_*.json files
    build_category_excel_report(args.out_dir, args.Split)

if __name__ == '__main__':
    main()